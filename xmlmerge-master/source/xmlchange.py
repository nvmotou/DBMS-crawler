# -*- coding:utf-8 -*-
# https://docs.python.org/2/library/xml.dom.html#dom-nodelist-objects
import xml.dom.minidom
import os.path
import time
from openpyxl import Workbook
start = time.clock()
wb=Workbook()
# sheet=wb.create_sheet(0,'爬虫抓取表')
sheet=wb.create_sheet('爬虫抓取表',0) # 2.3.5
i = 1
def get_xmlnode(node,name):
    return node.getElementsByTagName(name) if node else []

# def files1(rootdir):
# 	file = []
# 	for parent,dirnames,filenames in os.walk(rootdir):
# 		if parent == rootdir:
# 			for filename in filenames:
# 				if filename.endswith('.xml'):
# 					file.append(filename)
# 			return file
# 		else:
# 			pass

def files(rootdir):
    file = []
    for parent,dirnames,filenames in os.walk(rootdir):
        for filename in filenames:
            if filename.endswith('.xml'):
                file.append((parent+'/'+filename).replace('\\','/'))
    return file

def writedata(i,j,item):
    if i == 1:
        try:
            sheet.cell(row = i,column= j).value = item.tagName
            sheet.cell(row = i+1,column= j).value = item.firstChild.data.strip()
        except:
            sheet.cell(row = i,column= j).value = item.tagName
            attrtext = []
            for key in item.attributes.keys():
                attrvalue = item.attributes[key]
                attrtext.append(attrvalue.value)
            sheet.cell(row = i+1,column= j).value = ','.join(attrtext)
    else:
        try:
            sheet.cell(row = i+1,column= j).value = item.firstChild.data.strip()
        except:
            attrtext = []
            for key in item.attributes.keys():
                attrvalue = item.attributes[key]
                attrtext.append(attrvalue.value)
            sheet.cell(row = i+1,column= j).value = ','.join(attrtext)


def curitem(nodelist,v1,v2,v3,tag='item'):
    if nodelist:
        for node in nodelist:
            nodelist2 = node.getElementsByTagName(tag) if node else []
            islast = curitem(nodelist2,v1,v2,v3)
            if islast==1:
                items = node.childNodes
                j=4
                for item in items:
                    global i
                    writedata(i,j,item)
                    j=j+1
                sheet.cell(row=i+1,column=1).value = v1
                sheet.cell(row=i+1,column=2).value = v2
                sheet.cell(row=i+1,column=3).value = v3
                i=i+1
        return 2
    else:
        return 1

def begin():
    sangjin = '''
		-----------------------------------------
		| 欢迎使用XML批量转化EXCEL，适用于爬虫	|
		|                                       |
		| 使用方法:                             |
		| 1.新建files目录，与运行目录同一级     |
		|   将XML文件放在files目录中            |
		|   进入运行目录运行后缀为exe的文件     |
		| 2.XML是有格式要求的，会生成相关信息   |
		-----------------------------------------
		| 时间：2015年11月26日                  |
		| 新浪微博：一只尼玛                    |
		| 微信/QQ：569929309                    |
		-----------------------------------------
                    |实习已经结束，现在正式踏入社会|
                    |工具可处理几艘客软件生成的文件|
                    |如处理其他格式的XML请您自定义|
	'''
    print(sangjin)

begin()
print("开始处理......")
path = '../爬虫生成表.xlsx'
xmls = files('../files/')
erpath = []
en = 0 #错误数

sheet.cell(row=1,column=1).value = '线索ID'
sheet.cell(row=1,column=2).value = '网址路径'
sheet.cell(row=1,column=3).value = '爬取时间'
for filename in xmls:
	try:
		doc =xml.dom.minidom.parse(filename)
		root = doc.documentElement
		noderoot = get_xmlnode(root,'item')
		clueid = get_xmlnode(root,'clueid')
		v1=clueid[0].childNodes[0].data
		clueid1 = get_xmlnode(root,'uri')
		v2=clueid1[0].childNodes[0].data
		clueid2 = get_xmlnode(root,'createdate')
		v3=clueid2[0].childNodes[0].data
		curitem(noderoot,v1,v2,v3)
		print("处理成功："+filename)
	except Exception as e:
		raise
		en = en+1
		erpath.append(filename+"\n"+str(e))
		pass

wb.save(path)
total = len(xmls)
if erpath:
	print("-"*50)
	print("提取失败的文件:")
	print('\n'.join(erpath))
print("-"*50)
print('总共处理XML条数：'+str(total))
print('提取XML失败条数：'+str(en))
print('处理成功的XML条数：'+str(total-en))
print('生成记录数：'+str(i-1))
print("-"*50)
end = time.clock()
print("程序总共运行时间 : %.03f 秒" %(end-start))
print("处理结束........")
input()
